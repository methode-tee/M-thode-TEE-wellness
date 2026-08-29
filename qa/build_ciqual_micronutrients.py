#!/usr/bin/env python3
"""Transforme la feuille officielle ANSES CIQUAL 2025 en import long Supabase.

Le script localise les constituants par leur EN-TÊTE, pas par une position de colonne
fragile. Les valeurs qualifiées (<, traces, tirets, ND...) ne sont jamais converties
en nombres : le moteur préfère « non documenté » à une quantité inventée.
"""
import csv, re, sys, unicodedata
from pathlib import Path
from openpyxl import load_workbook

NUTRIENTS = {
    "calcium_mg": (["calcium"], "mg"),
    "iron_mg": (["fer"], "mg"),
    "iodine_ug": (["iode"], "µg"),
    "magnesium_mg": (["magnesium"], "mg"),
    "phosphorus_mg": (["phosphore"], "mg"),
    "potassium_mg": (["potassium"], "mg"),
    "selenium_ug": (["selenium"], "µg"),
    "zinc_mg": (["zinc"], "mg"),
    "vitamin_d_ug": (["vitamine d"], "µg"),
    "vitamin_e_mg": (["vitamine e"], "mg"),
    "vitamin_c_mg": (["vitamine c"], "mg"),
    "vitamin_b1_mg": (["vitamine b1", "thiamine"], "mg"),
    "vitamin_b2_mg": (["vitamine b2", "riboflavine"], "mg"),
    "vitamin_b3_mg": (["vitamine b3", "niacine"], "mg"),
    "vitamin_b6_mg": (["vitamine b6"], "mg"),
    "vitamin_b9_ug": (["vitamine b9", "folates totaux"], "µg"),
    "vitamin_b12_ug": (["vitamine b12"], "µg"),
}
OMEGA3_PARTS = [
    ["18:3", "alpha-linolen"],
    ["20:5", "epa"],
    ["22:6", "dha"],
]
EXACT = re.compile(r"^[+-]?\d+(?:[,.]\d+)?$")

def norm(value):
    text=unicodedata.normalize("NFKD",str(value or "")).encode("ascii","ignore").decode().lower()
    return re.sub(r"\s+"," ",text).strip()

def number(value):
    if isinstance(value, (int, float)): return float(value)
    text = str(value or "").strip().replace(" ", "")
    return float(text.replace(",", ".")) if EXACT.match(text) else None

def locate(headers, terms, unit=None):
    normalized=[norm(x) for x in headers]
    candidates=[]
    for i,h in enumerate(normalized):
        if any(norm(t) in h for t in terms):
            if unit and unit.lower()=="mg" and ("microg" in h or "µg" in str(headers[i]).lower()):
                continue
            candidates.append(i)
    return candidates[0] if candidates else None

def main(source, output):
    ws = load_workbook(source, read_only=True, data_only=True)["composition nutritionnelle"]
    headers=[cell.value for cell in next(ws.iter_rows(min_row=1,max_row=1))]
    code_idx=locate(headers,["alim_code","code aliment"]) or 6
    columns={}
    for key,(terms,unit) in NUTRIENTS.items():
        idx=locate(headers,terms,unit)
        if idx is not None: columns[key]=(idx,unit)
    omega_cols=[]
    for terms in OMEGA3_PARTS:
        idx=locate(headers,terms)
        if idx is not None: omega_cols.append(idx)
    output = Path(output); output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["ciqual_code","nutrient_key","value_100g","unit","source","source_version"])
        writer.writeheader()
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[code_idx] or "").strip()
            if not code: continue
            values = {}
            for key, (idx, unit) in columns.items():
                value = number(row[idx]) if idx < len(row) else None
                if value is not None and value >= 0: values[key] = (value, unit)
            omega3 = [number(row[idx]) for idx in omega_cols if idx < len(row)]
            if omega3 and any(value is not None for value in omega3): values["omega3_g"] = (sum(value or 0 for value in omega3), "g")
            for key, (value, unit) in values.items():
                writer.writerow({"ciqual_code":code,"nutrient_key":key,"value_100g":format(value,".10g"),"unit":unit,"source":"ANSES - Table Ciqual 2025","source_version":"2025-11-03"})
    print("Colonnes CIQUAL détectées:", ", ".join(sorted(columns)), "+ omega3" if omega_cols else "")

if __name__ == "__main__":
    if len(sys.argv) != 3: raise SystemExit("usage: build_ciqual_micronutrients.py SOURCE.xlsx OUTPUT.csv")
    main(sys.argv[1], sys.argv[2])
